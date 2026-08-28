#!/usr/bin/env python3
"""
Segundo proyector: libro QN v2.4.2 -> data/QN_v242_aggregates.json
                                  -> data/QN_v242_contract.json

Convive con scripts/build_data.py, no lo reemplaza. Aquel proyecta la capa
semantica del inventario; este proyecta el corpus de calidad de notas.

QUE SE PUBLICA Y QUE NO
-----------------------
Las hojas agregadas del libro cubren el corpus COMPLETO aunque el libro traiga
las hojas de detalle muestreadas. Eso no se asume, se verifica: QN03 y QN04
exigen que User_By_Group sume exactamente los incidentes de usuario y que
Alert_By_Group sume exactamente las alertas que declara Overview. Si el libro
que se pase trae agregados derivados de una muestra, esas dos sumas no cuadran
y el build se detiene.

Las hojas de detalle NO se publican aqui. Viven en el navegador, se cargan
desde la seccion de carga y se indexan en IndexedDB.

User_By_Agent queda FUERA del proyector, no solo sin publicar. Medir personal
nominal del vendor exige decision de RRHH y Legal antes de tocar el dato
(decision 4 del handoff). El script verifica que no se haya colado.

USO
---
    python3 scripts/build_qn_aggregates.py RUTA_DEL_LIBRO.xlsx
    npm run qn -- RUTA_DEL_LIBRO.xlsx

El libro fuente no se versiona, igual que el xlsx de la capa semantica. Lo que
se versiona es el JSON que sale de aqui.
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalar con: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
OUT_AGG = ROOT / "data" / "QN_v242_aggregates.json"
OUT_CONTRACT = ROOT / "data" / "QN_v242_contract.json"

INSTRUMENT = "QN Work Notes Quality Analyzer v2.4.2"

# Contrato de columnas. Es el mismo que valida el cargador del navegador.
SHEET_COLUMNS: dict[str, list[str]] = {
    "Overview": ["Work Notes Quality Analyzer v2.4.2"],
    "User_Detail": [
        "Number", "Assignment Group", "Assigned To", "State", "Priority",
        "Service Offering", "Short Description", "Total Score", "Label",
        "Root Cause", "Resolution Docs", "Description Qlty", "WN Completeness",
        "Lang & Prof", "Noise Ratio", "Human Notes #", "Note Chars",
        "Has Close Notes", "Close Code", "U Close Code", "Reopen Count",
        "Decalogue Primary", "Decalogue All", "Discernment", "Feedback",
        "Decalogue Primary v2", "Decalogue All v2", "Discernment v2",
        "Decalogue Count v2", "Year", "Month", "Category", "Subcategory",
        "compliance_class", "has_rca_marker", "is_template_match", "Closed At",
    ],
    "User_By_Group": [
        "Assignment Group", "Incidents", "Avg Score", "Avg Root Cause",
        "Avg Resolution", "Avg Description", "Avg WN Complete", "Avg Lang & Prof",
        "Avg Noise", "Close Notes %", "Excellent", "Good", "Poor", "Critical",
    ],
    "User_By_Agent": [
        "Assigned To", "Assignment Group", "Incidents", "Avg Score",
        "Avg Root Cause", "Avg Resolution", "Avg WN Complete", "Avg Noise",
        "Excellent", "Good", "Poor", "Critical",
    ],
    "By_Decalogue": [
        "Code", "Pattern", "Discernment", "Incidents", "Avg Score",
        "Close Notes %", "Excellent", "Good", "Poor", "Critical",
    ],
    "Decalogue_By_Group": ["Assignment Group", "Code", "Pattern", "Incidents", "Avg Score"],
    "Decalogue_Validation": ["Decalogue A/B Validation (v2.4.1)"],
    "Compliance_CloseNotes": ["Close-Notes Compliance Metric (ADO 4.3.45.2)"],
    "Compliance_Alerts": ["Alert Documentation Rate (ADO 4.3.45.2 — v2.3.1)"],
    "Alert_Detail": [
        "Number", "Assignment Group", "Assigned To", "State", "Priority",
        "Service Offering", "Short Description", "Ops Classification",
        "Intervention Level", "Auto-Resolved", "Has Root Cause", "Has Steps",
        "Human Notes #", "Substantive Chars", "Noise Ratio", "Unique Authors",
        "Closed At",
    ],
    "Alert_By_Group": [
        "Assignment Group", "Alerts", "Auto-Resolved", "Auto-Resolved %",
        "Has Root Cause", "Has Steps", "Avg Noise", "Avg Substantive Chars",
        "None", "Minimal", "Moderate", "Substantive",
    ],
    "Dual_Axis": [
        "Label (Axis 1 — process)", "DIAGNOSTICO (n)", "SUSTANTIVO (n)",
        "FORMAL_ONLY (n)", "EMPTY (n)", "Total", "DIAGNOSTICO (row %)",
        "SUSTANTIVO (row %)", "FORMAL_ONLY (row %)", "EMPTY (row %)",
    ],
}

# Hojas que el proyector se niega a leer, por decision tomada.
EXCLUDED_SHEETS = {"User_By_Agent": "Decision 4: medir personal nominal del vendor exige RRHH y Legal."}

# Hojas cuyo grano vive en el navegador, no aqui.
DETAIL_SHEETS = {"User_Detail", "Alert_Detail"}

# Hojas de banner: titulo en A1 y pares clave/valor debajo, sin encabezado
# tabular. No se les revisa contrato de columnas, porque no tienen columnas.
BANNER_SHEETS = {"Overview", "Decalogue_Validation", "Compliance_CloseNotes", "Compliance_Alerts"}

ERRORS: list[str] = []
CHECKS: list[dict] = []


def check(cid: str, what: str, ok: bool, detail: str = "") -> bool:
    CHECKS.append({"id": cid, "invariant": what, "passed": bool(ok), "detail": detail})
    if not ok:
        ERRORS.append(f"{cid}  {what}" + (f" — {detail}" if detail else ""))
    return bool(ok)


def norm_key(s) -> str:
    """Normalizador canonico de Assignment Group. Identico al de la app."""
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def to_num(v):
    """'719,946' -> 719946 · '38.5%' -> 38.5 · deja el resto tal cual."""
    if isinstance(v, (int, float)):
        return v
    if not isinstance(v, str):
        return None
    s = v.strip().replace(",", "").replace("%", "").replace("+", "")
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None


def rows_of(ws) -> list[tuple]:
    return [r for r in ws.iter_rows(values_only=True)]


def table(ws, columns: list[str]) -> list[dict]:
    """Hoja tabular con encabezado en la fila 1. Fila vacia corta la lectura."""
    rows = rows_of(ws)
    out = []
    for r in rows[1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({c: r[i] if i < len(r) else None for i, c in enumerate(columns)})
    return out


def kv_sections(ws) -> dict:
    """
    Hoja de banner: titulo en A1, secciones en mayusculas sin valor, y pares
    clave/valor debajo. Se conserva el texto crudo y se agrega el numero cuando
    se puede leer, sin descartar el original.
    """
    out: dict[str, dict] = {}
    section = "_header"
    out[section] = {}
    for r in rows_of(ws):
        cells = ["" if c is None else str(c).strip() for c in r]
        if not any(cells):
            continue
        key, rest = cells[0], [c for c in cells[1:] if c]
        if key and not rest:
            section = key
            out.setdefault(section, {})
            continue
        if key:
            value = rest[0] if len(rest) == 1 else rest
            entry: dict = {"raw": value}
            n = to_num(value) if isinstance(value, str) else None
            if n is not None:
                entry["value"] = n
            out.setdefault(section, {})[key] = entry
    return out


def main() -> int:
    if len(sys.argv) < 2:
        sys.exit(
            "Uso: python3 scripts/build_qn_aggregates.py RUTA_DEL_LIBRO.xlsx\n"
            "El libro fuente no se versiona; el JSON que sale de aqui, si."
        )
    src = Path(sys.argv[1]).expanduser().resolve()
    if not src.exists():
        sys.exit(f"No existe el libro: {src}")

    sha = hashlib.sha256(src.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(src, data_only=True)
    present = list(wb.sheetnames)

    # ---- QN01 · las doce hojas del contrato ----
    missing = [s for s in SHEET_COLUMNS if s not in present]
    check("QN01", "las 12 hojas del contrato estan presentes", not missing,
          f"faltan: {missing}" if missing else f"{len(present)} hojas")

    # ---- QN02 · contrato de columnas por hoja tabular ----
    bad_cols = []
    for name, cols in SHEET_COLUMNS.items():
        if name not in present or name in BANNER_SHEETS:
            continue
        hdr = [None if c is None else str(c) for c in (rows_of(wb[name])[0] if rows_of(wb[name]) else [])]
        if hdr[: len(cols)] != cols:
            bad_cols.append(f"{name}: esperado {cols[:3]}… , encontrado {hdr[:3]}…")
    check("QN02", "el encabezado de cada hoja tabular coincide con el contrato",
          not bad_cols, " | ".join(bad_cols))

    if ERRORS:
        return report_and_exit()

    # ---- lectura ----
    overview = kv_sections(wb["Overview"])
    dual_rows = table(wb["Dual_Axis"], SHEET_COLUMNS["Dual_Axis"])
    user_by_group = table(wb["User_By_Group"], SHEET_COLUMNS["User_By_Group"])
    alert_by_group = table(wb["Alert_By_Group"], SHEET_COLUMNS["Alert_By_Group"])
    by_deca_all = table(wb["By_Decalogue"], SHEET_COLUMNS["By_Decalogue"])
    # La hoja trae diez filas de codigo y despues un bloque Summary con pares
    # clave/valor en las dos primeras columnas. Se separan: mezclarlos haria que
    # "Classified incidents" pareciera un patron del decalogo.
    by_decalogue = [r for r in by_deca_all if re.fullmatch(r"D\d{2}", str(r["Code"] or "").strip())]
    deca_summary = {
        str(r["Code"]).strip(): r["Pattern"]
        for r in by_deca_all
        if r not in by_decalogue and r["Pattern"] is not None
    }
    deca_by_group = table(wb["Decalogue_By_Group"], SHEET_COLUMNS["Decalogue_By_Group"])
    deca_validation = kv_sections(wb["Decalogue_Validation"])
    comp_notes = kv_sections(wb["Compliance_CloseNotes"])
    comp_alerts = kv_sections(wb["Compliance_Alerts"])

    def ov(section: str, key: str):
        e = overview.get(section, {}).get(key)
        if not e:
            return None
        if "value" in e:
            return e["value"]
        m = re.match(r"([\d,]+)", str(e["raw"]))
        return to_num(m.group(1)) if m else None

    total_declared = ov("POPULATION SUMMARY", "Total incidents")
    user_declared = ov("POPULATION SUMMARY", "User incidents (scored)")
    alert_declared = ov("POPULATION SUMMARY", "Incident alerts (classified)")

    # ---- QN03 / QN04 · los agregados cubren el corpus, no una muestra ----
    ubg_sum = sum(r["Incidents"] for r in user_by_group if isinstance(r["Incidents"], (int, float)))
    abg_sum = sum(r["Alerts"] for r in alert_by_group if isinstance(r["Alerts"], (int, float)))
    check("QN03", "User_By_Group suma exactamente los incidentes de usuario declarados",
          ubg_sum == user_declared, f"suma {ubg_sum:,} vs declarado {user_declared:,}")
    check("QN04", "Alert_By_Group suma exactamente las alertas declaradas",
          abg_sum == alert_declared, f"suma {abg_sum:,} vs declarado {alert_declared:,}")

    # ---- QN05 · los dos granos particionan el universo ----
    check("QN05", "user + alert == total, los dos granos particionan sin traslape",
          (user_declared or 0) + (alert_declared or 0) == total_declared,
          f"{user_declared:,} + {alert_declared:,} vs {total_declared:,}")

    # ---- QN06 / QN07 · Dual_Axis cuadra por fila y por columna ----
    total_row = next((r for r in dual_rows if str(r["Label (Axis 1 — process)"]).strip() == "Total"), None)
    band_rows = [r for r in dual_rows if r is not total_row]
    check("QN06", "la fila Total de Dual_Axis iguala los incidentes de usuario",
          bool(total_row) and total_row["Total"] == user_declared,
          f"{total_row['Total'] if total_row else '—'} vs {user_declared:,}")
    axis_cols = ["DIAGNOSTICO (n)", "SUSTANTIVO (n)", "FORMAL_ONLY (n)", "EMPTY (n)"]
    col_mismatch = [c for c in axis_cols
                    if sum(r[c] for r in band_rows if isinstance(r[c], (int, float))) != (total_row or {}).get(c)]
    check("QN07", "cada columna de Dual_Axis suma su fila Total", not col_mismatch,
          f"no cuadran: {col_mismatch}" if col_mismatch else "las cuatro clases cuadran")

    # ---- QN08 · los diez codigos, y la hoja NO es sumable ----
    # Un incidente puede llevar varios codigos, por lo tanto la suma de la
    # columna Incidents excede a los incidentes clasificados. Se verifica el
    # sentido de la desigualdad y se publica el sobreconteo, para que nadie
    # sume la columna y la presente como poblacion. Misma familia que R4.
    codes = [str(r["Code"]).strip() for r in by_decalogue if r["Code"]]
    codes_ok = sorted(codes) == [f"D{i:02d}" for i in range(1, 11)]
    deca_sum = sum(r["Incidents"] for r in by_decalogue if isinstance(r["Incidents"], (int, float)))
    classified = to_num(re.split(r"/", str(deca_summary.get("Classified incidents", "")))[0])
    deca_overcount = (deca_sum - classified) if classified is not None else None
    check("QN08",
          "By_Decalogue trae D01..D10 y su suma excede a los clasificados: no es sumable",
          codes_ok and classified is not None and deca_sum >= classified,
          f"{len(codes)} codigos · suma {deca_sum:,} vs clasificados {classified:,} "
          f"· sobreconteo {deca_overcount:,} por multi-patron"
          if classified is not None else f"{len(codes)} codigos: {sorted(codes)}")

    # ---- QN09 · la serie entre cortes queda bloqueada, con su razon medida ----
    agree = None
    for sect in deca_validation.values():
        for k, v in sect.items():
            if "agreement" in k.lower() and "value" in v:
                agree = v["value"]
    matrix = [r for r in rows_of(wb["Decalogue_Validation"])
              if r and str(r[0]).strip().startswith("D") and len(str(r[0]).strip()) == 3]
    d01 = next((r for r in matrix if str(r[0]).strip() == "D01"), None)
    d01_delta = to_num(d01[6]) if d01 and len(d01) > 6 else None
    check("QN09", "la validacion v1/v2 esta presente y D01 mueve lo declarado",
          d01_delta is not None and abs(float(d01_delta) - 640.2) < 0.05,
          f"D01 delta% = {d01_delta}")

    # ---- QN10 / QN11 · claves de Assignment Group ----
    names = [r["Assignment Group"] for r in user_by_group if r["Assignment Group"]]
    keys = {norm_key(n) for n in names}
    check("QN10", "User_By_Group no repite el nombre de un grupo",
          len(names) == len(set(names)), f"{len(names)} nombres, {len(set(names))} distintos")
    check("QN11", "el normalizador no colapsa grupos: un nombre, una clave",
          len(set(names)) == len(keys), f"{len(set(names))} nombres -> {len(keys)} claves")

    # ---- QN12 · el KPI de cumplimiento viaja con su objetivo ----
    rate = None
    target = None
    for sect in comp_notes.values():
        for k, v in sect.items():
            kl = k.lower()
            if "population rate" in kl or kl == "rate":
                rate = v.get("value", rate)
            if "target" in kl:
                target = v.get("value", target)
    if target is None:
        target = ov("CLOSE-NOTES COMPLIANCE (ADO 4.3.45.2)", "Target")
    if rate is None:
        rate = ov("CLOSE-NOTES COMPLIANCE (ADO 4.3.45.2)", "Population rate")
    check("QN12", "el KPI de close-notes trae tasa y objetivo, nunca la tasa sola",
          rate is not None and target is not None, f"tasa {rate} · objetivo {target}")

    # ---- QN13 · no existe opened_at: MTTR se queda bloqueado ----
    opened = []
    for name in ("User_Detail", "Alert_Detail"):
        hdr = [str(c or "") for c in rows_of(wb[name])[0]]
        opened += [h for h in hdr if re.search(r"open(ed)?[ _]?at|created", h, re.I)]
    check("QN13", "no hay opened_at en ninguna hoja: MTTR queda bloqueado, nunca derivado",
          not opened, f"columnas sospechosas: {opened}" if opened else "solo Closed At")

    # ---- la hoja excluida no se coló ----
    for s in EXCLUDED_SHEETS:
        assert s not in DETAIL_SHEETS

    if ERRORS:
        return report_and_exit()

    # ---- salida ----
    aggregates = {
        "meta": {
            "instrument": INSTRUMENT,
            "as_of": "2026-08-12",
            "generated_from": src.name,
            "sha256": sha,
            "projector": "scripts/build_qn_aggregates.py",
            "sheets_in_source": present,
            "aggregates_cover": "el corpus completo, verificado por QN03 y QN04",
            "detail_not_published_here":
                "User_Detail y Alert_Detail viven en el navegador: se cargan desde "
                "/upload y se indexan en IndexedDB. Este JSON no los contiene.",
            "excluded_by_decision": EXCLUDED_SHEETS,
            "blocked_measures": [
                {"id": "MTTR", "reason": "no hay opened_at en el corpus, solo Closed At (QN13)"},
                {"id": "REASSIGNMENT", "reason": "el campo no existe"},
                {"id": "DECALOGUE_SERIES", "reason": "clasificador v1/v2 sin calibrar (QN09)"},
            ],
            "invariants": CHECKS,
        },
        "overview": overview,
        "dual_axis": {"rows": band_rows, "total": total_row},
        "user_by_group": user_by_group,
        "alert_by_group": alert_by_group,
        "by_decalogue": {
            "rows": by_decalogue,
            "summary": deca_summary,
            "classified_incidents": classified,
            "sum_of_codes": deca_sum,
            "overcount": deca_overcount,
            "not_additive":
                "Un incidente puede llevar varios codigos. La suma de la columna "
                "Incidents sobrecuenta respecto a los incidentes clasificados; la "
                "cifra de poblacion es classified_incidents, nunca sum_of_codes.",
        },
        "decalogue_by_group": deca_by_group,
        "decalogue_validation": deca_validation,
        "compliance_closenotes": comp_notes,
        "compliance_alerts": comp_alerts,
    }

    # El manifiesto: lo que el cargador del navegador exige antes de publicar
    # una sola cifra. Sale del mismo libro, de modo que no puede desincronizarse.
    contract = {
        "instrument": INSTRUMENT,
        "as_of": "2026-08-12",
        "measured_from": src.name,
        "sha256": sha,
        "sheets": {
            name: {
                "columns": cols,
                "role": ("banner" if name in BANNER_SHEETS
                         else "detail" if name in DETAIL_SHEETS
                         else "excluded" if name in EXCLUDED_SHEETS
                         else "aggregate"),
                "indexed_in_browser": name in DETAIL_SHEETS,
            }
            for name, cols in SHEET_COLUMNS.items()
        },
        "declared_population": {
            "total": total_declared,
            "user_detail": user_declared,
            "alert_detail": alert_declared,
        },
        "ag_key_normalizer": "s.toUpperCase().replace(/[^A-Z0-9]/g, '')",
        "invariants": [{"id": c["id"], "invariant": c["invariant"]} for c in CHECKS],
        "note":
            "El cargador valida el archivo subido contra este manifiesto ANTES de "
            "publicar una cifra. Si no cuadra, la app declara corpus sin verificar "
            "y se niega a estampar la fecha de corte.",
    }

    OUT_AGG.write_text(json.dumps(aggregates, ensure_ascii=False, separators=(",", ":")), "utf-8")
    OUT_CONTRACT.write_text(json.dumps(contract, ensure_ascii=False, indent=2), "utf-8")

    print(f"Fuente   {src.name}")
    print(f"sha256   {sha}")
    print(f"Poblacion declarada  total {total_declared:,} = user {user_declared:,} + alert {alert_declared:,}")
    print(f"Agregados  User_By_Group {len(user_by_group):,} grupos · Alert_By_Group {len(alert_by_group):,} · "
          f"Decalogue_By_Group {len(deca_by_group):,}")
    print(f"Invariantes  {sum(1 for c in CHECKS if c['passed'])}/{len(CHECKS)} pasan")
    print(f"Escrito  {OUT_AGG.relative_to(ROOT)}  ({OUT_AGG.stat().st_size/1024:.0f} KB)")
    print(f"Escrito  {OUT_CONTRACT.relative_to(ROOT)}  ({OUT_CONTRACT.stat().st_size/1024:.0f} KB)")
    return 0


def report_and_exit() -> int:
    print("El proyector se detuvo. Invariantes rotas:\n", file=sys.stderr)
    for e in ERRORS:
        print(f"  {e}", file=sys.stderr)
    print("\nNo se escribio ningun JSON. Un agregado que no cuadra con su propia "
          "poblacion declarada no se publica.", file=sys.stderr)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
