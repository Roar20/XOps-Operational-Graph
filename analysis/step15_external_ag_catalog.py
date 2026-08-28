#!/usr/bin/env python3
"""
Step 1.5 addendum 2 - the external Assignment Groups catalogue (Drive, Apr 2026).

analysis/source_drop/Assignment_Groups_Catalog_v2026_230426.xlsx carries two
things the loaded corpus does not:

  sheet Mapping_BA_AG_BIOps - BusinessApplication -> Assignment Group from a
    BIOps outreach campaign. Independent provenance from BOTH sources the
    projector uses, so for the first time in this project an Application -> AG
    claim can be checked against an outside opinion.
  sheet AGs - a DECLARED classification of 280 assignment groups into
    AI/ML | BI | Platform/Infra, with a per-row confidence.

It does NOT carry an incident number, a per-incident Business Application, or a
Configuration Item. It is application-grain. Ground truth for
Incident -> Application remains NOT AVAILABLE and nothing here changes that.

Requires openpyxl.  Read-only.
"""
import json, re, collections, difflib, os, sys
try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

SRC = "analysis/source_drop"
XL  = os.path.join(SRC, "Assignment_Groups_Catalog_v2026_230426.xlsx")
canon = lambda s: re.sub(r"[^A-Z0-9]", "", (s or "").upper())

wb   = openpyxl.load_workbook(XL, data_only=True)
R    = json.load(open("data/xops-operational-graph-data.json"))
E    = json.load(open(os.path.join(SRC, "edges_app_ag.json")))
AGC  = {g["ag_id"]: g for g in json.load(open(os.path.join(SRC, "catalog_assignment_groups.json")))}
APPC = {a["app_id"]: a for a in json.load(open(os.path.join(SRC, "catalog_applications.json")))}

MODEL = {canon(a["name"]): a for a in R["applications"]}
FULL  = collections.defaultdict(set)
for e in E:
    FULL[canon(APPC[e["app_id"]]["name"])].add(AGC[e["ag_id"]]["ag_key"])

print("=" * 74)
print("1. WHAT THIS FILE IS, AND IS NOT")
print("=" * 74)
print(f"  sheets: {wb.sheetnames}")
mrows = [r for r in list(wb["Mapping_BA_AG_BIOps"].iter_rows(values_only=True))[1:] if r[0]]
hdr = [c for c in list(wb["Mapping_BA_AG_BIOps"].iter_rows(values_only=True))[0]]
print(f"  Mapping_BA_AG_BIOps columns: {hdr}")
print("\n  No incident number. No per-incident Business Application. No")
print("  Configuration Item. Grain is one row per application-group pair.")
print("  GROUND TRUTH for Incident -> Application remains NOT AVAILABLE.")

ext = collections.defaultdict(set)
for r in mrows:
    ext[canon(r[0])].add(canon(r[1]))
matched = [k for k in ext if k in MODEL]
print(f"\n  rows {len(mrows)} | distinct applications {len(ext)} | "
      f"matching the 504 portfolio {len(matched)} ({100*len(matched)/len(ext):.1f}%)")

print("\n" + "=" * 74)
print("2. EVIDENCE CONVERGENCE - Application -> Assignment Group")
print("=" * 74)
print("  First convergence measurement available to this project. Agreement is")
print("  defined as a NON-EMPTY INTERSECTION between the external group set and")
print("  the model's. That is a weak test where an application carries several")
print("  groups, and it is reported as such - it is not an accuracy.")

def run(getset, tag):
    ag = cf = sil = 0; conf = []
    for k in matched:
        mine = getset(k)
        if not mine:
            sil += 1; continue
        if ext[k] & mine: ag += 1
        else:
            cf += 1
            conf.append((MODEL[k]["name"], sorted(mine), sorted(ext[k])))
    tot = ag + cf + sil
    print(f"\n  {tag}")
    print(f"    CONVERGENCE {ag:>4}  ({100*ag/tot:.1f}%)")
    print(f"    CONFLICT    {cf:>4}  ({100*cf/tot:.1f}%)")
    print(f"    model silent{sil:>4}  (model declares no group, external names one)")
    return conf

c_pub  = run(lambda k: {canon(g) for g in MODEL[k]["ags"]}, "published model  (629 edges)")
c_full = run(lambda k: FULL.get(k, set()),                  "full edge set    (672 edges)")

print(f"\n  The 43 edges the projector discards account for "
      f"{len(c_pub)-len(c_full)} of the {len(c_pub)} conflicts.")
print("  This is the first measurement that gives the build_data.py:152-158")
print("  finding an operational cost: restoring the discarded edges lowers the")
print("  conflict rate against an independent source.")

print("\n  Remaining conflicts, split by whether the two names are near-identical")
print("  (an entity-resolution artefact) or genuinely different groups:")
genuine = 0
for n, m, e in sorted(c_full):
    best = max((difflib.SequenceMatcher(None, x, y).ratio() for x in m for y in e), default=0)
    kind = "NAME VARIANT" if best >= 0.80 else "GENUINE"
    if kind == "GENUINE": genuine += 1
    print(f"    [{kind:<12}] {n[:36]:<38} model={m[:2]}")
    print(f"                   {'':<38} ext  ={e[:2]}")
print(f"\n  Genuine disagreements: {genuine} of {len(matched)} matched applications "
      f"({100*genuine/len(matched):.1f}%).")
print("  Convergence is NOT correctness. Two sources can agree and both be wrong,")
print("  and this measures a DECLARED relationship, not an incident attribution.")

print("\n" + "=" * 74)
print("3. ENTITY SPACE - declared classification vs the report's heuristic")
print("=" * 74)
arows = [r for r in list(wb["AGs"].iter_rows(values_only=True))[1:] if r[0]]
print(f"  {len(arows)} assignment groups, declared Category:")
for k, v in collections.Counter(r[1] for r in arows).most_common():
    print(f"    {k:<18}{v:>5}")
print("  per-row Confidence:")
for k, v in collections.Counter(r[2] for r in arows).most_common():
    print(f"    {k:<18}{v:>5}")
mk = {g["ag_key"] for g in R["assignment_groups"]}
ek = {canon(r[0]) for r in arows}
print(f"\n  AG keys: external {len(ek)} | model {len(mk)} | overlap {len(ek & mk)} "
      f"| external-only {len(ek - mk)} | model-only {len(mk - ek)}")
print("\n  This REPLACES the lexicon heuristic in section D of the report for the")
print(f"  {len(ek & mk)} overlapping groups - but only for the CONFIRMED rows. The")
print("  catalogue marks 81 of its own rows INFERRED and 35 PENDING; those carry")
print("  no more authority than the heuristic they would replace and must not be")
print("  promoted. The model's own AG namespace is only "
      f"{100*len(ek & mk)/len(mk):.0f}% covered by this catalogue.")
